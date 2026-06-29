"""Excel workbook builder (Summary + per-task sheets).

Ported from the reference harness ``_write_workbook`` and friends. Adapted to
return bytes (for artifact-service upload) instead of writing to disk, and to
take the frontend base URL as a parameter instead of reading global settings.
The reference's dead ``Insights`` sheet is intentionally omitted.
"""

from __future__ import annotations

import logging
import math
from io import BytesIO
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.reports.models import IterationRecord, MODEL_ORDER
from app.reports.snapshot import extract_record_model_response
from app.reports.summary import extract_record_status, select_prompt, select_prompt_id

logger = logging.getLogger(__name__)


def _sanitize_for_excel(text: str) -> str:
    """Sanitize text for Excel: replace problematic unicode, strip control chars, truncate."""
    if not text:
        return ""

    replacements = {
        "•": "-", "✅": "[OK]", "❌": "[FAIL]", "⚠️": "[WARN]", "🔍": "[SEARCH]",
        "📝": "[NOTE]", "🎯": "[TARGET]", "🚀": "[LAUNCH]", "💡": "[IDEA]", "⭐": "[STAR]",
        "🔥": "[HOT]", "💯": "[100]", "📊": "[CHART]", "📈": "[UP]", "📉": "[DOWN]",
        "🔧": "[TOOL]", "⚡": "[FAST]", "🎉": "[CELEBRATE]", "🏆": "[TROPHY]", "📋": "[CLIPBOARD]",
        "🔒": "[LOCK]", "🔓": "[UNLOCK]", "📌": "[PIN]", "📍": "[LOCATION]", "⏰": "[TIME]",
        "📅": "[CALENDAR]", "📧": "[EMAIL]", "📞": "[PHONE]", "🌐": "[WEB]", "💻": "[COMPUTER]",
        "📱": "[MOBILE]", "💾": "[SAVE]", "📁": "[FOLDER]", "📄": "[DOCUMENT]", "🔗": "[LINK]",
    }

    sanitized = text
    for unicode_char, ascii_replacement in replacements.items():
        sanitized = sanitized.replace(unicode_char, ascii_replacement)

    # Remove remaining control characters (except newlines and tabs)
    sanitized = "".join(char for char in sanitized if ord(char) >= 32 or char in "\n\t")

    # Truncate if too long (Excel cell limit is 32,767 characters)
    if len(sanitized) > 32000:
        sanitized = sanitized[:32000] + "... [TRUNCATED]"

    return sanitized


def _safe_sheet_name(workbook: Workbook, task_id: str) -> str:
    invalid_chars = [":", "\\", "/", "?", "*", "[", "]"]
    sanitized = task_id
    for char in invalid_chars:
        sanitized = sanitized.replace(char, "_")

    base = sanitized[:28] if len(sanitized) > 28 else sanitized
    candidate = base or "Task"
    suffix = 1
    existing = {ws.title for ws in workbook.worksheets}
    while candidate in existing:
        candidate = f"{base[:25]}_{suffix}" if len(base) > 25 else f"{base}_{suffix}"
        suffix += 1
    return candidate


def _get_difficulty_colors(difficulty: str) -> Dict[str, str]:
    if difficulty == "hard":
        return {"bg": "FFC7CE", "text": "000000", "border": "FFC7CE"}
    elif difficulty == "medium":
        return {"bg": "FFEB9C", "text": "000000", "border": "FFEB9C"}
    elif difficulty == "easy":
        return {"bg": "C6EFCE", "text": "000000", "border": "C6EFCE"}
    else:
        return {"bg": "FFFFFF", "text": "000000", "border": "FFFFFF"}


def _get_model_colors(runner_key: str) -> Dict[str, str]:
    runner_lower = runner_key.lower()
    if "gpt" in runner_lower or "openai" in runner_lower:
        return {"header_bg": "1976D2", "header_text": "FFFFFF", "data_bg": "E3F2FD", "data_text": "000000"}
    elif "claude" in runner_lower or "anthropic" in runner_lower:
        return {"header_bg": "7B1FA2", "header_text": "FFFFFF", "data_bg": "F3E5F5", "data_text": "000000"}
    elif "gemini" in runner_lower or "google" in runner_lower:
        return {"header_bg": "5D4037", "header_text": "FFFFFF", "data_bg": "EFEBE9", "data_text": "000000"}
    else:
        return {"header_bg": "424242", "header_text": "FFFFFF", "data_bg": "F5F5F5", "data_text": "000000"}


def _populate_summary_sheet(ws, rows: List[Dict[str, object]], total_iterations: Optional[int] = None, *, include_steps: bool = False) -> None:
    model_columns: List[Tuple[str, str, str]] = []
    if rows:
        for model_key, model_label in MODEL_ORDER:
            column_name = f"{model_label} Breaking"
            if any(column_name in row for row in rows):
                model_columns.append((model_key, model_label, column_name))

    headers = ["Prompt ID", "Prompt"]
    for _, model_label, _ in model_columns:
        headers.append(f"{model_label} Breaking")
        if include_steps:
            headers.append(f"{model_label} Median Steps")

    ws.append(headers)

    for cell in ws[1]:
        cell.fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(rows, 2):
        row_data: List[object] = [
            _sanitize_for_excel(str(row.get("Prompt ID", ""))),
            _sanitize_for_excel(str(row.get("Prompt", ""))),
        ]
        for _, model_label, _ in model_columns:
            row_data.append(_sanitize_for_excel(str(row.get(f"{model_label} Breaking", ""))))
            if include_steps:
                median_steps = row.get(f"{model_label} Median Steps")
                if median_steps is not None:
                    try:
                        row_data.append(int(median_steps))
                    except (ValueError, TypeError):
                        row_data.append(0)
                else:
                    row_data.append(0)

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="center", vertical="center")

            if col_idx > 2:
                col_offset = col_idx - 3
                if include_steps:
                    is_breaking_column = (col_offset % 2 == 0)
                    model_idx = col_offset // 2
                else:
                    is_breaking_column = True
                    model_idx = col_offset

                if is_breaking_column and model_idx < len(model_columns):
                    _, model_label, _ = model_columns[model_idx]
                    model_difficulty = row.get(f"{model_label} Difficulty", "unknown")
                    model_difficulty_colors = _get_difficulty_colors(model_difficulty)
                    if str(value).strip() and str(value).strip() not in ["No data", ""]:
                        cell.fill = PatternFill(start_color=model_difficulty_colors["bg"], end_color=model_difficulty_colors["bg"], fill_type="solid")
                        cell.font = Font(color=model_difficulty_colors["text"])
                    else:
                        cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                        cell.font = Font(color="000000")

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    def _apply_widths() -> None:
        col = 1
        ws.column_dimensions[get_column_letter(col)].width = 28  # Prompt ID
        col += 1
        ws.column_dimensions[get_column_letter(col)].width = 120  # Prompt
        col += 1
        for _ in model_columns:
            ws.column_dimensions[get_column_letter(col)].width = 35
            col += 1
            if include_steps:
                ws.column_dimensions[get_column_letter(col)].width = 35
                col += 1

    _apply_widths()

    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = wrap_alignment
        ws.row_dimensions[row[0].row].height = 80

    definitions_start_row = ws.max_row + 3
    _add_difficulty_definitions_table(ws, definitions_start_row, total_iterations, include_steps=include_steps)

    # Re-apply widths (the definitions table may override them)
    _apply_widths()

    if include_steps:
        _add_median_steps_chart(ws, rows, model_columns, definitions_start_row)


def _add_median_steps_chart(ws, rows: List[Dict[str, object]], model_columns: List[Tuple[str, str, str]], definitions_start_row: int) -> None:
    try:
        from openpyxl.chart import BarChart, Reference
        from openpyxl.chart.layout import Layout, ManualLayout

        if not rows or not model_columns:
            return

        has_valid_data = False
        for row in rows:
            for _, model_label, _ in model_columns:
                median_steps = row.get(f"{model_label} Median Steps")
                if median_steps is not None and median_steps != 0:
                    has_valid_data = True
                    break
            if has_valid_data:
                break

        if not has_valid_data:
            return

        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        chart.title = "Median Steps by Task and Model"

        num_rows = len(rows)
        chart.height = 15
        chart.width = 30
        chart.y_axis.title = "Median Steps"
        chart.x_axis.title = "Task ID"
        chart.y_axis.majorGridlines = None
        chart.x_axis.majorGridlines = None

        max_steps = 0
        for row in rows:
            for _, model_label, _ in model_columns:
                median_steps = row.get(f"{model_label} Median Steps")
                if median_steps is not None:
                    max_steps = max(max_steps, int(median_steps))

        if max_steps > 0:
            chart.y_axis.scaling.min = 0
            chart.y_axis.scaling.max = int(max_steps * 1.2)

        chart.x_axis.delete = False
        chart.y_axis.delete = False
        chart.x_axis.tickLblSkip = 1

        for idx, (_, model_label, _) in enumerate(model_columns):
            col_num = 4 + (idx * 2)
            data = Reference(ws, min_col=col_num, min_row=1, max_row=num_rows + 1)
            chart.add_data(data, titles_from_data=True)

        cats = Reference(ws, min_col=1, min_row=2, max_row=num_rows + 1)
        chart.set_categories(cats)
        chart.legend.position = "tr"
        chart.legend.overlay = False
        chart.layout = Layout(manualLayout=ManualLayout(x=0.1, y=0.1, w=0.85, h=0.55))

        ws.add_chart(chart, f"C{definitions_start_row + 8}")
    except Exception as e:  # pragma: no cover - chart rendering is best-effort
        logger.error(f"Chart error: {e}")


def _add_difficulty_definitions_table(ws, start_row: int, total_iterations: Optional[int] = None, *, include_steps: bool = False) -> None:
    title_cell = ws.cell(row=start_row, column=3, value="Difficulty Definitions")
    title_cell.font = Font(bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=start_row, start_column=3, end_row=start_row, end_column=5)

    header_row = start_row + 1
    headers = ["Difficulty", "Criteria", "Description"]
    for col_idx, header in enumerate(headers, 3):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D0D0D0", end_color="D0D0D0", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    N = int(total_iterations) if isinstance(total_iterations, int) and total_iterations > 0 else 8

    pass_min_medium = math.floor(0.4 * N) + 1
    pass_max_medium = max(N - 1, 0)
    fail_min_medium = 1 if N > 1 else 0
    fail_max_medium = max(N - pass_min_medium, 0)
    pass_max_hard = math.floor(0.4 * N)
    fail_min_hard = math.ceil(0.6 * N)

    difficulty_data: List[Dict[str, str]] = []

    if include_steps:
        criteria = f"Pass@{N} = 100% AND Median Steps ≤ 25"
        if N == 1:
            description = "Prompt should pass the single iteration AND Median Steps should be ≤ 25"
        else:
            description = f"Prompt should pass {N} times (0 failures) AND Median Steps should be ≤ 25"
    else:
        criteria = f"Pass@{N} = 100%"
        if N == 1:
            description = "Prompt should pass the single iteration"
        else:
            description = f"Prompt should fail 0 times and pass {N} times"
    difficulty_data.append({"difficulty": "Easy", "criteria": criteria, "description": description, "fill_color": "C6EFCE"})

    if include_steps:
        if N == 1:
            criteria = "(Pass@1 < 100% AND Pass@1 ≥ 50%) OR (25 < Median Steps ≤ 50)"
            description = "Prompt fails single iteration but could pass OR Median Steps between 26-50"
        elif N == 2:
            criteria = f"(50% ≤ Pass@{N} < 100%) OR (25 < Median Steps ≤ 50)"
            description = "Prompt should pass 1 out of 2 times OR Median Steps between 26-50"
        else:
            criteria = f"(40% < Pass@{N} < 100%) OR (25 < Median Steps ≤ 50)"
            description = (
                f"Prompt should pass between {pass_min_medium}-{pass_max_medium} out of {N} times OR "
                f"Median Steps between 26-50"
            )
        difficulty_data.append({"difficulty": "Medium", "criteria": criteria, "description": description, "fill_color": "FFEB9C"})
    else:
        if N >= 2:
            if N == 2:
                criteria = f"Pass@{N} ≥ 50%"
                description = "Prompt should pass 1 out of 2 times and fail 1 out of 2 times"
            else:
                criteria = f"40% < Pass@{N} < 100%"
                description = (
                    f"Prompt should fail between {fail_min_medium}-{fail_max_medium} out of {N} times and "
                    f"pass between {pass_min_medium}-{pass_max_medium} out of {N} times"
                )
            difficulty_data.append({"difficulty": "Medium", "criteria": criteria, "description": description, "fill_color": "FFEB9C"})

    if include_steps:
        if N == 1:
            criteria = "Pass@1 < 100% OR Median Steps > 50"
            description = "Prompt should fail the single iteration OR Median Steps should be > 50"
        elif N == 2:
            criteria = "Pass@2 < 50% OR Median Steps > 50"
            description = "Prompt should fail both iterations OR Median Steps should be > 50"
        else:
            criteria = f"Pass@{N} < 40% OR Median Steps > 50"
            description = f"Prompt should pass between 0-{pass_max_hard} out of {N} times OR Median Steps should be > 50"
    else:
        if N == 1:
            criteria = "Pass@1 < 100%"
            description = "Prompt should fail the single iteration and pass 0 times"
        elif N == 2:
            criteria = f"Pass@{N} < 40%"
            description = "Prompt should fail both iterations and pass 0 times"
        else:
            criteria = f"Pass@{N} < 40%"
            description = (
                f"Prompt should fail between {fail_min_hard}-{N} out of {N} times and "
                f"pass between 0-{pass_max_hard} out of {N} times"
            )
    difficulty_data.append({"difficulty": "Hard", "criteria": criteria, "description": description, "fill_color": "FFC7CE"})

    for row_idx, data in enumerate(difficulty_data, header_row + 1):
        difficulty_cell = ws.cell(row=row_idx, column=3, value=data["difficulty"])
        difficulty_cell.fill = PatternFill(start_color=data["fill_color"], end_color=data["fill_color"], fill_type="solid")
        difficulty_cell.alignment = Alignment(horizontal="center", vertical="center")

        criteria_cell = ws.cell(row=row_idx, column=4, value=data["criteria"])
        criteria_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        description_cell = ws.cell(row=row_idx, column=5, value=data["description"])
        description_cell.alignment = Alignment(wrap_text=True, vertical="top")

        ws.row_dimensions[row_idx].height = 60

    ws.column_dimensions[get_column_letter(3)].width = 20
    ws.column_dimensions[get_column_letter(4)].width = 40
    ws.column_dimensions[get_column_letter(5)].width = 40


def _populate_task_sheet(ws, task_id: str, runner_map: Dict[str, List[IterationRecord]], base_execution_url: str) -> None:
    has_steps_data = False
    for records in runner_map.values():
        for record in records:
            if record.total_steps is not None:
                has_steps_data = True
                break
        if has_steps_data:
            break

    model_column_headers = ["Model Response", "Pass/Fail", "Time Taken"]
    if has_steps_data:
        model_column_headers.append("Steps")
    model_column_headers.extend(["Iteration Link", "Comments"])

    top_headers = ["Prompt ID", "Prompt"]
    second_headers = ["Prompt ID", "Prompt"]

    models_with_data: List[Tuple[str, str]] = []
    for runner_key, label in MODEL_ORDER:
        if runner_key in runner_map and runner_map[runner_key]:
            models_with_data.append((runner_key, label))
            top_headers.extend([label] * len(model_column_headers))
            second_headers.extend(model_column_headers)

    ws.append(top_headers)
    ws.append(second_headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
    for cell in ws[2]:
        cell.font = Font(bold=True)

    current_col = 3
    for runner_key, label in models_with_data:
        colors = _get_model_colors(runner_key)
        start_col = current_col
        end_col = start_col + len(model_column_headers) - 1
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = PatternFill(start_color=colors["header_bg"], end_color=colors["header_bg"], fill_type="solid")
            cell.font = Font(bold=True, color=colors["header_text"])
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=2, column=col)
            cell.fill = PatternFill(start_color=colors["header_bg"], end_color=colors["header_bg"], fill_type="solid")
            cell.font = Font(bold=True, color=colors["header_text"])
        current_col = end_col + 1

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

    current_col = 3
    for _, label in models_with_data:
        start_col = current_col
        end_col = start_col + len(model_column_headers) - 1
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        ws.cell(row=1, column=start_col).value = label
        current_col = end_col + 1

    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center")

    runner_records: Dict[str, List[IterationRecord]] = {}
    max_runs = 0
    for runner_key, records in runner_map.items():
        records_sorted = sorted(
            records,
            key=lambda record: (
                record.iteration if record.iteration is not None else float("inf"),
                record.start_timestamp or "",
                record.end_timestamp or "",
                record.run_id or "",
            ),
        )
        runner_records[runner_key] = records_sorted
        max_runs = max(max_runs, len(records_sorted))

    if max_runs == 0:
        max_runs = 1

    for row_index in range(max_runs):
        row_records: List[Optional[IterationRecord]] = []
        for runner_key, _label in models_with_data:
            records = runner_records.get(runner_key, [])
            row_records.append(records[row_index] if row_index < len(records) else None)

        if all(record is None for record in row_records):
            continue

        prompt_value = select_prompt(runner_map) or ""
        prompt_id_value = select_prompt_id(runner_map) or task_id
        is_primary_row = row_index == 0
        sanitized_prompt = _sanitize_for_excel(prompt_value) if is_primary_row else ""

        row: List[object] = [
            str(prompt_id_value or task_id) if is_primary_row else "",
            sanitized_prompt,
        ]

        hyperlink_targets: List[Tuple[int, str, str]] = []

        for record in row_records:
            if record:
                status = extract_record_status(record) or ""
                iteration_uuid = getattr(record, "iteration_uuid", None)
                auto_comments = (getattr(record, "verification_comments", "") or "") if iteration_uuid else ""

                model_response = extract_record_model_response(record) or ""
                if not model_response:
                    model_response = "—"
                model_response = _sanitize_for_excel(model_response)

                status_label = _sanitize_for_excel(status)
                comments_label = _sanitize_for_excel(auto_comments) if auto_comments else "—"
                execution_uuid = getattr(record, "execution_uuid", None)
                hyperlink_display = "—"
                hyperlink_url = None
                if execution_uuid and iteration_uuid:
                    iteration_uuid_str = str(iteration_uuid)
                    hyperlink_url = f"{base_execution_url}/executions/{execution_uuid}/iterations/{iteration_uuid_str}"
                    hyperlink_display = f"View Iteration {iteration_uuid_str[:8]}"
            else:
                auto_comments = ""
                model_response = ""
                status_label = ""
                comments_label = ""
                hyperlink_display = "—"
                hyperlink_url = None

            time_taken = ""
            if record and record.duration_seconds is not None:
                minutes = record.duration_seconds / 60.0
                if minutes < 1:
                    time_taken = f"{record.duration_seconds:.1f}s"
                else:
                    time_taken = f"{minutes:.1f}m"
            elif record and record.timelapse:
                time_taken = record.timelapse

            steps_value = ""
            if has_steps_data and record and record.total_steps is not None:
                steps_value = str(record.total_steps)

            row_len_before = len(row)
            row_data: List[object] = [model_response, status_label, time_taken]
            if has_steps_data:
                row_data.append(steps_value)
            row_data.extend([hyperlink_display, auto_comments])
            row.extend(row_data)
            if hyperlink_url:
                hyperlink_col_offset = 4 if has_steps_data else 3
                hyperlink_targets.append((row_len_before + hyperlink_col_offset + 1, hyperlink_url, hyperlink_display))

        ws.append(row)

        excel_row = ws.max_row
        current_col = 3
        for runner_key, _label in models_with_data:
            colors = _get_model_colors(runner_key)
            start_col = current_col
            end_col = start_col + len(model_column_headers) - 1
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=excel_row, column=col)
                cell.fill = PatternFill(start_color=colors["data_bg"], end_color=colors["data_bg"], fill_type="solid")
                cell.font = Font(color=colors["data_text"])
            current_col = end_col + 1

        if hyperlink_targets:
            for col_idx, url, display in hyperlink_targets:
                cell = ws.cell(row=excel_row, column=col_idx)
                cell.value = display
                if url and url.startswith(("http://", "https://")):
                    cell.hyperlink = url
                    cell.font = Font(color="0563C1", underline="single")
                else:
                    cell.value = display
                    cell.font = Font(color="000000")

    column_widths = [28, 120]
    if has_steps_data:
        model_widths = [80, 26, 30, 20, 36, 60]
    else:
        model_widths = [80, 26, 30, 36, 60]
    for _ in models_with_data:
        column_widths.extend(model_widths)

    for idx, width in enumerate(column_widths):
        ws.column_dimensions[get_column_letter(idx + 1)].width = width

    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    for row_idx in range(3, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 160
        for cell in ws[row_idx]:
            cell.alignment = wrap_alignment


def build_workbook_bytes(
    summary_rows: List[Dict[str, object]],
    iterations: List[IterationRecord],
    task_rows: Dict[str, Dict[str, List[IterationRecord]]],
    *,
    total_iterations: Optional[int] = None,
    frontend_base_url: str = "",
) -> bytes:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"

    include_steps = any(r.total_steps is not None for r in (iterations or []))
    _populate_summary_sheet(summary_ws, summary_rows, total_iterations, include_steps=include_steps)

    for task_id, runner_map in sorted(task_rows.items()):
        sheet_name = _safe_sheet_name(wb, task_id)
        task_ws = wb.create_sheet(sheet_name)
        _populate_task_sheet(task_ws, task_id, runner_map, frontend_base_url)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
