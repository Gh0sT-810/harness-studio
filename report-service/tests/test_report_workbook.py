from io import BytesIO

from openpyxl import load_workbook

from app.reports.summary import build_summary
from app.reports.workbook import build_workbook_bytes


def test_workbook_sheets_summary_and_per_task_no_insights(make_record):
    records = [
        make_record(task_id="T1", runner="anthropic", iteration=1, status="passed", execution_uuid="e1", iteration_uuid="i1", total_steps=10),
        make_record(task_id="T1", runner="openai", iteration=1, status="failed", execution_uuid="e2", iteration_uuid="i2", total_steps=40),
        make_record(task_id="T2", runner="anthropic", iteration=1, status="passed", execution_uuid="e3", iteration_uuid="i3", total_steps=12),
    ]
    summary_rows, task_map = build_summary(records)
    data = build_workbook_bytes(summary_rows, records, task_map, total_iterations=8, frontend_base_url="http://localhost:3000")

    wb = load_workbook(BytesIO(data))
    assert "Summary" in wb.sheetnames
    assert "Insights" not in wb.sheetnames
    assert "T1" in wb.sheetnames
    assert "T2" in wb.sheetnames


def test_summary_sheet_has_dynamic_breaking_columns(make_record):
    records = [
        make_record(task_id="T1", runner="anthropic", iteration=1, status="passed"),
        make_record(task_id="T1", runner="openai", iteration=1, status="failed"),
    ]
    summary_rows, task_map = build_summary(records)
    data = build_workbook_bytes(summary_rows, records, task_map, total_iterations=8, frontend_base_url="http://x")

    wb = load_workbook(BytesIO(data))
    header = [cell.value for cell in wb["Summary"][1]]
    assert "Claude Sonnet 4 Breaking" in header
    assert "OpenAI Computer Use Preview Breaking" in header
    # gemini has no data → no column
    assert "Google Gemini Computer Use Breaking" not in header


def test_task_sheet_has_iteration_hyperlink(make_record):
    records = [make_record(task_id="T1", runner="anthropic", iteration=1, status="passed", execution_uuid="e1", iteration_uuid="i1")]
    summary_rows, task_map = build_summary(records)
    data = build_workbook_bytes(summary_rows, records, task_map, frontend_base_url="http://localhost:3000")

    wb = load_workbook(BytesIO(data))
    task_ws = wb["T1"]
    links = [cell.hyperlink.target for row in task_ws.iter_rows() for cell in row if cell.hyperlink]
    assert any("/executions/e1/iterations/i1" in link for link in links)


def test_workbook_without_steps_omits_steps_column(make_record):
    records = [make_record(task_id="T1", runner="anthropic", iteration=1, status="passed", total_steps=None)]
    summary_rows, task_map = build_summary(records)
    data = build_workbook_bytes(summary_rows, records, task_map, frontend_base_url="http://x")

    wb = load_workbook(BytesIO(data))
    second_header = [cell.value for cell in wb["T1"][2]]
    assert "Steps" not in second_header
    assert "Model Response" in second_header
